"""Tolerance / robustness analysis — adaptive multi-parameter Monte Carlo.

Key features
------------
- **Latin Hypercube Sampling** (not pure random) for better space coverage.
- **Adaptive densification**: starts with ``min_samples``, adds ``batch_size``
  per round until output statistics converge or ``max_samples`` is reached.
- **Two-step solve**: calibrates f_data to the true resonant frequency before
  each measurement pass, ensuring field monitors are sampled at resonance.
- **Incremental Excel**: writes results after every evaluation — no data loss
  on interruption.

tolerance_abs definition
------------------------
``tolerance_abs`` = full ± manufacturing tolerance band.
σ = tolerance_abs / 3, samples transformed via Normal inverse CDF from
a uniform LHS design.

Usage
-----
::

    .venv\\Scripts\\python examples\\tolerance_analysis.py
"""

import sys
import os as _os
_PROJECT_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_SRC_DIR = _os.path.join(_PROJECT_ROOT, "src")
if _SRC_DIR not in sys.path:
    sys.path.insert(0, _SRC_DIR)

import time, yaml
from datetime import datetime
import numpy as np
from scipy.stats import norm, qmc

from cst_optimization.core.connection import CSTConnection
from cst_optimization.core.project import CSTProject
from cst_optimization.core.results import ResultReader, ResultBundle
from cst_optimization.core.solver import SolverRunner
from cst_optimization.physics.cavity import (
    ResonantFrequency, LoadedQ, CouplingBeta, IntrinsicQ,
    PeakSurfaceField, InputPower, MinS11,
)
from cst_optimization.utils.units import GHz, MV_per_m
from cst_optimization.core.messages import MessageLogger
from cst_optimization.physics.poynting import max_modified_poynting, discover_field_files
from cst_optimization.physics.heating import max_h_from_field_file, pulsed_heating_delta_t
from cst_optimization.core.retry import EvaluationRetryHandler, RetryConfig, RetryTier
from cst_optimization.workflows.recovery import EvaluationResult, EvaluationStatus

CONFIG_PATH = _os.path.join(_PROJECT_ROOT, "config", "default.yaml")


# ── Sampling ────────────────────────────────────────────────────────────


def generate_lhs_normal_samples(
    nominal: np.ndarray, sigma: np.ndarray,
    n_samples: int, seed: int,
) -> np.ndarray:
    """LHS in [0,1]^D transformed to Normal(nominal, sigma).

    Latin Hypercube ensures better coverage than pure random sampling
    — one sample per "slice" of each parameter's distribution.
    """
    n_dims = len(nominal)
    sampler = qmc.LatinHypercube(d=n_dims, seed=seed)
    unit = sampler.random(n=n_samples)  # (N, D) in [0, 1]
    # Transform uniform → Normal via inverse CDF, then scale
    normal = norm.ppf(unit)  # N(0, 1)
    samples = nominal + normal * sigma
    # Clip at ±5σ
    for j in range(n_dims):
        lo = nominal[j] - 5 * sigma[j]
        hi = nominal[j] + 5 * sigma[j]
        samples[:, j] = np.clip(samples[:, j], lo, hi)
    return samples


# ── Two-step solve helpers ──────────────────────────────────────────────


def _calibration_solve(project, param_names, x, f_data_guess, runner, msg_logger, sample_idx=0):
    """Pass 1: solve → return (f0_GHz_or_None, reason, com_retry).

    CST 2026 may report "Terminated abnormally" even when results exist.
    We only trust "com" errors as definitive — for everything else we try
    to save and read results anyway.
    """
    import time as _t
    reason = ""
    try:
        params = dict(zip(param_names, x))
        params["f_data"] = f_data_guess
        project.update_parameters(params)
        _t.sleep(10.0)

        result = runner.run(project)
        msg_logger.capture(project)

        if not result.success:
            if result.error_type == "com":
                reason = "COM connection lost"
                return None, reason, True
            # mesh / timeout / unclassified — solver may still have results
            reason = f"solver reported {result.error_type or 'error'}, trying to read anyway"

        # Try save + read regardless of solver return code
        try:
            project.save()
        except Exception:
            pass

        try:
            project.model3d.abort_solver()
            _t.sleep(1.0)
        except Exception:
            pass

        reader = ResultReader(project.filename, allow_interactive=True)
        s11 = reader.get_s_parameter()
        mag = np.abs(s11.s_complex)
        from cst_optimization.physics.formulas import half_power_bandwidth
        f0, _, _, _ = half_power_bandwidth(s11.frequencies, mag, target_freq=f_data_guess)

        if np.isfinite(f0) and f0 > 0:
            return float(f0), "", False
        else:
            return None, "f0 invalid after read", False

    except Exception as exc:
        return None, str(exc)[:80], False


def _measurement_solve(project, param_names, x, f0_ghz, runner, msg_logger, sample_idx=0):
    """Pass 2: solve with corrected f_data = f0 → return (result_dict, com_retry).

    Same tolerance as calibration: solver reports of "Terminated abnormally"
    don't prevent us from reading results that were actually produced.
    """
    import time as _t
    result = {
        "solver_ok": False, "error": "", "f0_ghz": f0_ghz,
        "q_loaded": np.nan, "coupling_beta": np.nan, "q0": np.nan,
        "e_peak": np.nan, "s11_db": np.nan, "p_input_mw": np.nan,
        "Sc_max": np.nan, "DeltaT_K": np.nan, "field_flatness": np.nan,
    }
    try:
        params = dict(zip(param_names, x))
        params["f_data"] = f0_ghz
        project.update_parameters(params)
        _t.sleep(10.0)

        solver_result = runner.run(project)
        msg_logger.capture(project)

        if not solver_result.success:
            if solver_result.error_type == "com":
                result["error"] = "COM connection lost"
                return result, True
            # mesh / timeout / unclassified — may still have readable results
            result["error"] = f"solver reported {solver_result.error_type or 'error'}, trying to read anyway"

        # Try save + read regardless of solver return code
        try:
            project.save()
        except Exception:
            pass

        # Poll for field exports
        prj_dir = _os.path.splitext(project.filename)[0]
        for _ in range(3):
            _t.sleep(5.0)
            e_file, _ = discover_field_files(prj_dir)
            if e_file:
                break

        reader = ResultReader(project.filename, allow_interactive=True)
        s11 = reader.get_s_parameter()
        e_max = reader.get_scalar(reader.TREEPATH_MAX_E_Z0)
        bundle = ResultBundle(s_parameters={"S1,1": s11}, scalars={"MaxE_Z0": e_max})
        result["solver_ok"] = True
        result["q_loaded"] = LoadedQ().compute(bundle)
        result["coupling_beta"] = CouplingBeta().compute(bundle)
        result["q0"] = IntrinsicQ().compute(bundle)
        result["e_peak"] = PeakSurfaceField().compute(bundle)
        result["s11_db"] = 20 * np.log10(max(MinS11().compute(bundle), 1e-15))
        result["p_input_mw"] = InputPower(target_e_acc_vm=200*MV_per_m).compute(bundle) / 1e6

        _cache_field_exports(project.filename, sample_idx)
        _compute_field_outputs(result, project.filename, e_max.value, sample_idx,
                               reader)

    except Exception as exc:
        result["error"] = str(exc)[:120]
    return result, False


def _compute_field_outputs(result: dict, prj_path: str, e_sim: float,
                           sample_idx: int, reader) -> None:
    """Compute Sc_max, DeltaT_K, field_flatness from cached exports & 0D templates."""
    # --- Sc_max (modified Poynting vector, scaled to 200 MV/m) ---
    try:
        prj_dir = _os.path.splitext(prj_path)[0]
        cache_dir = _os.path.join(_os.path.dirname(prj_dir), "Results", "fields",
                                  f"sample_{sample_idx:04d}")
        e_file, h_file = discover_field_files(cache_dir)
        if not e_file:
            e_file, h_file = discover_field_files(prj_dir)

        if e_file and h_file and e_sim > 0:
            scale = 200e6 / e_sim
            result["Sc_max"] = max_modified_poynting(e_file, h_file, gc=0.125,
                                                     field_scale=scale)
    except Exception:
        pass

    # --- DeltaT_K (pulsed heating) ---
    try:
        prj_dir = _os.path.splitext(prj_path)[0]
        cache_dir = _os.path.join(_os.path.dirname(prj_dir), "Results", "fields",
                                  f"sample_{sample_idx:04d}")
        _, h_file = discover_field_files(cache_dir)
        if not h_file:
            _, h_file = discover_field_files(prj_dir)
        if h_file and e_sim > 0:
            h_peak = max_h_from_field_file(h_file)
            result["DeltaT_K"] = pulsed_heating_delta_t(
                h_peak_sim=h_peak, e_peak_sim=e_sim,
                e_target=200e6, pulse_width_ns=300,
                frequency_hz=11.424e9, rrr=5.5,
            )
    except Exception:
        pass

    # --- field_flatness ---
    try:
        e1 = reader.get_scalar(reader.TREEPATH_MAX_E_Z1).value
        e2 = reader.get_scalar(reader.TREEPATH_MAX_E_Z2).value
        e_max_val = max(e_sim, e1, e2)
        e_min_val = min(e_sim, e1, e2)
        if e_max_val > 0:
            result["field_flatness"] = 1.0 - e_min_val / e_max_val
    except Exception:
        pass


def _cache_field_exports(prj_path: str, sample_idx: int) -> None:
    """Copy E/H field exports from project Export/3d/ to a cache directory."""
    import shutil, glob as _glob
    prj_dir = _os.path.splitext(prj_path)[0]
    src_dir = _os.path.join(prj_dir, "Export", "3d")
    if not _os.path.isdir(src_dir):
        return
    dst_dir = _os.path.join(_os.path.dirname(prj_dir), "Results", "fields", f"sample_{sample_idx:04d}")
    _os.makedirs(dst_dir, exist_ok=True)
    for pat in ("*e-field*", "*E-field*", "*E_Field*", "*h-field*", "*H-field*", "*H_Field*"):
        for f in _glob.glob(_os.path.join(src_dir, pat)):
            shutil.copy2(f, _os.path.join(dst_dir, _os.path.basename(f)))


# ── Convergence check ───────────────────────────────────────────────────


def _check_convergence(
    prev_means: dict[str, float],
    current_means: dict[str, float],
    rtol: float,
) -> bool:
    """Return True if ALL outputs have converged (relative change < rtol)."""
    if not prev_means:
        return False
    for key in current_means:
        if key not in prev_means:
            return False
        old, new = prev_means[key], current_means[key]
        if abs(old) < 1e-15:
            continue  # skip near-zero outputs
        if abs(new - old) / abs(old) > rtol:
            return False
    return True


# ── Incremental Excel ───────────────────────────────────────────────────


def _save_checkpoint(filepath, completed_indices, failed_permanent=None):
    """Save completed and permanently-failed sample indices to a JSON file."""
    import json as _json
    data = {
        "completed": sorted(completed_indices),
        "failed_permanent": sorted(failed_permanent or set()),
    }
    with open(filepath, "w") as fh:
        _json.dump(data, fh)


def _excel_init(filepath, headers):
    """Create Excel file with headers if it doesn't exist yet.

    Does NOT overwrite an existing file — watchdog restarts will
    resume appending to the same workbook.
    """
    if _os.path.isfile(filepath):
        return  # file already exists, keep existing data
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Tolerance"
    ws.append(headers); ws.freeze_panes = "A2"
    _os.makedirs(_os.path.dirname(filepath) or ".", exist_ok=True)
    wb.save(filepath)


def _excel_append(filepath, row_data):
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    wb.active.append(row_data)
    wb.save(filepath)


# ── Main ────────────────────────────────────────────────────────────────


def main() -> None:
    cfg = yaml.safe_load(open(CONFIG_PATH, "r", encoding="utf-8"))
    tc = cfg.get("tolerance", {})
    if not tc:
        print("No 'tolerance' section."); return

    min_samples = tc.get("min_samples", 30)
    max_samples = tc.get("max_samples", 100)
    batch_size = tc.get("batch_size", 10)
    conv_rtol = tc.get("convergence_rtol", 0.01)
    seed = tc.get("seed", 42)
    project_path = tc["project_path"]

    param_entries = [p for p in tc.get("parameters", []) if p.get("enabled")]
    output_entries = [o for o in tc.get("outputs", []) if o.get("enabled")]
    if not param_entries:
        print("No parameters enabled."); return
    if not output_entries:
        print("No outputs enabled."); return

    nominal = np.array([p["nominal"] for p in param_entries])
    sigma = np.array([p["tolerance_abs"] / 3.0 for p in param_entries])
    param_names = [p["name"] for p in param_entries]
    n_params = len(param_entries)

    print(f"Parameters ({n_params}):")
    for j in range(n_params):
        print(f"  {param_names[j]:15s} nom={nominal[j]:10.4f}  ±{sigma[j]*3:.4f} (±3σ) {param_entries[j].get('unit','')}")
    print(f"\nOutputs ({len(output_entries)}):")
    for o in output_entries:
        print(f"  {o['name']:20s}  {o.get('description','')}")

    # ── Excel headers ─────────────────────────────────────────────────
    log_dir = cfg.get("logging", {}).get("output_dir", "D:/Results")
    xlsx_path = _os.path.join(log_dir, "tolerance_analysis.xlsx")
    headers = (
        ["#", "batch"] + list(param_names) + ["f0_calibrated_ghz"]
        + [o["name"] for o in output_entries]
        + ["solver_ok", "elapsed_s", "error", "timestamp"]
    )
    _excel_init(xlsx_path, headers)

    # ── Adaptive loop ─────────────────────────────────────────────────
    solver_cfg = cfg.get("solver", {})
    runner = SolverRunner(
        timeout_s=solver_cfg.get("stagnation_timeout_s", 300),
        settle_s=solver_cfg.get("settle_s", 2.0),
    )

    # ── Checkpoint / resume ──────────────────────────────────────────
    checkpoint_path = xlsx_path.replace(".xlsx", "_checkpoint.json")
    completed_indices: set = set()
    failed_permanent: set = set()

    if _os.path.exists(checkpoint_path):
        import json as _json
        with open(checkpoint_path, "r") as fh:
            ck = _json.load(fh)
        completed_indices = set(ck.get("completed", []))
        failed_permanent = set(ck.get("failed_permanent", []))
        print(f"Checkpoint loaded: {len(completed_indices)} completed, "
              f"{len(failed_permanent)} failed-permanent, resuming...")
    elif _os.path.exists(xlsx_path):
        # Auto-import from existing Excel (previous run without checkpoint)
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    idx = int(row[0]) - 1  # Excel row #1 → index 0
                    if idx >= 0:
                        completed_indices.add(idx)
            wb.close()
            print(f"Auto-imported {len(completed_indices)} results from existing Excel")
            _save_checkpoint(checkpoint_path, completed_indices, failed_permanent)
        except Exception:
            pass

    # CST message logger for diagnostics
    msg_log_dir = _os.path.join(log_dir, "messages")
    msg_logger = MessageLogger(output_dir=msg_log_dir, enabled=True)

    all_results: list[dict] = []
    prev_means: dict[str, float] = {}
    eval_count = 0
    batch = 0

    # One-time connection — reconnect only on COM errors
    conn = CSTConnection(cfg["cst"]["library_path"], mode="any_or_new")
    conn.connect()
    conn.set_quiet_mode(True)
    project = conn.open_project(project_path)
    print(f"\nConnected to CST (PID: {conn.pid})")

    # ── Retry handler setup ───────────────────────────────────────────
    retry_cfg_raw = tc.get("retry", {})
    retry_handler = None
    _project_ref = project
    _conn_ref = conn

    if retry_cfg_raw.get("enabled", True):
        retry_config = RetryConfig(
            enabled=True,
            max_tier1=int(retry_cfg_raw.get("max_tier1", 3)),
            max_tier2=int(retry_cfg_raw.get("max_tier2", 2)),
            evaluation_timeout_s=float(retry_cfg_raw.get("evaluation_timeout_s", 600.0)),
        )

        # on_reconnect callback — update connection & invalidate project
        def _on_reconnect(new_conn):
            nonlocal _conn_ref, _project_ref
            _conn_ref = new_conn
            if _project_ref is not None:
                try:
                    _project_ref.close(save=False)
                except Exception:
                    pass
                _project_ref = None

            # Re-open project with the new connection
            _project_ref = new_conn.open_project(project_path)

        # Evaluate closure: calibrate + measure as a single retryable unit
        def evaluate(params, iteration):
            nonlocal _project_ref, _conn_ref
            x = np.asarray(params, dtype=float)

            # Pass 1: calibration
            f0_ghz, cal_reason, com_retry = _calibration_solve(
                _project_ref, param_names, x, 11.424, runner, msg_logger, sample_idx=iteration,
            )
            if com_retry:
                return EvaluationResult(
                    status=EvaluationStatus.COM_LOST,
                    error=f"Calibration COM lost: {cal_reason}",
                )
            if f0_ghz is None or not np.isfinite(f0_ghz):
                if "shallow" in str(cal_reason).lower() or "invalid" in str(cal_reason).lower():
                    return EvaluationResult(
                        status=EvaluationStatus.PHYSICS_INVALID,
                        error=f"Calibration: {cal_reason}",
                    )
                return EvaluationResult(
                    status=EvaluationStatus.SOLVER_FAILED,
                    error=f"Calibration: {cal_reason}",
                )

            # Pass 2: measurement
            result_dict, com_retry = _measurement_solve(
                _project_ref, param_names, x, f0_ghz, runner, msg_logger, sample_idx=iteration,
            )
            if com_retry:
                return EvaluationResult(
                    status=EvaluationStatus.COM_LOST,
                    error="Measurement COM lost",
                    f0_ghz=f0_ghz,
                )

            raw_metrics = {"f0_calibrated_ghz": f0_ghz}
            for out_name in ["q_loaded", "coupling_beta", "q0", "e_peak",
                             "s11_db", "p_input_mw", "Sc_max", "DeltaT_K",
                             "field_flatness"]:
                raw_metrics[out_name] = result_dict.get(out_name, np.nan)

            status = (EvaluationStatus.SUCCESS if result_dict.get("solver_ok")
                      else EvaluationStatus.SOLVER_FAILED)
            return EvaluationResult(
                status=status,
                f0_ghz=f0_ghz,
                raw_metrics=raw_metrics,
                error=result_dict.get("error", ""),
            )

        retry_handler = EvaluationRetryHandler(
            connection=conn,
            project_path=project_path,
            library_path=cfg["cst"]["library_path"],
            config=retry_config,
            on_reconnect=_on_reconnect,
        )
        print(f"Retry handler: Tier1={retry_config.max_tier1}, Tier2={retry_config.max_tier2}, "
              f"timeout={retry_config.evaluation_timeout_s}s")

    current_n = min_samples
    while current_n <= max_samples:
            batch += 1
            batch_seed = seed + batch * 1000
            samples = generate_lhs_normal_samples(nominal, sigma, current_n, seed=batch_seed)

            start_idx = eval_count
            n_new = current_n - start_idx
            n_skip = len([i for i in range(start_idx, current_n)
                         if i in completed_indices or i in failed_permanent])
            print(f"\n{'─'*60}")
            print(f"  BATCH {batch}: {n_new} samples (total {current_n})"
                  + (f" — skipping {n_skip} already done" if n_skip else ""))
            print(f"{'─'*60}")

            for i in range(start_idx, current_n):
                if i in completed_indices:
                    print(f"  [{i+1}/{current_n}] SKIP (already done)")
                    eval_count += 1
                    continue
                if i in failed_permanent:
                    print(f"  [{i+1}/{current_n}] SKIP (failed-permanent)")
                    eval_count += 1
                    continue

                x = samples[i]
                vals_str = ", ".join(
                    f"{n}={x[j]:.4f}" for j, n in enumerate(param_names)
                )
                print(f"  [{i+1}/{current_n}] {vals_str}")

                t0 = time.perf_counter()

                if retry_handler is not None:
                    # ── Three-tier retry path ─────────────────────────
                    eval_result, tier = retry_handler.execute(evaluate, x, i)
                    elapsed = time.perf_counter() - t0

                    # Map EvaluationResult back to result dict
                    result = {
                        "solver_ok": eval_result.solver_ok,
                        "error": eval_result.error,
                        "x": x,
                        "f0_calibrated_ghz": eval_result.f0_ghz,
                        "f0_ghz": eval_result.f0_ghz,
                        "elapsed_s": elapsed,
                    }
                    if eval_result.raw_metrics:
                        for k, v in eval_result.raw_metrics.items():
                            result[k] = v

                    if tier != RetryTier.TIER1:
                        print(f"    [retry tier: {tier.name}]", flush=True)

                    if eval_result.solver_ok:
                        print(f"OK ({elapsed:.0f}s)  Beta={result.get('coupling_beta',np.nan):.3f}")
                    elif eval_result.status == EvaluationStatus.PHYSICS_INVALID:
                        print(f"FAIL (physics) [{eval_result.error[:60]}]")
                    else:
                        print(f"FAIL ({elapsed:.0f}s) [{eval_result.error[:60]}]")

                    if tier == RetryTier.EXHAUSTED:
                        failed_permanent.add(i)
                else:
                    # ── Legacy inline COM retry path ─────────────────
                    print(f"    Calibration ... ", end="", flush=True)
                    f0_ghz, cal_reason, com_retry = _calibration_solve(
                        project, param_names, x, 11.424, runner, msg_logger, sample_idx=i
                    )

                    if com_retry:
                        print(f"COM error, reconnecting...", flush=True)
                        try:
                            conn.reconnect()
                            project = conn.open_project(project_path)
                            f0_ghz, cal_reason, _ = _calibration_solve(
                                project, param_names, x, 11.424, runner, msg_logger, sample_idx=i
                            )
                        except Exception:
                            f0_ghz = None
                            cal_reason = "COM unrecoverable"

                    if f0_ghz is None or not np.isfinite(f0_ghz):
                        print(f"FAIL  [{cal_reason[:60]}]")
                        result = {"solver_ok": False, "error": f"f0 calibration: {cal_reason}", "x": x}
                        _excel_append(xlsx_path, _make_row(i, batch, param_names, output_entries, result))
                        all_results.append(result); eval_count += 1
                        completed_indices.add(i)
                        _save_checkpoint(checkpoint_path, completed_indices, failed_permanent)
                        continue
                    print(f"f0={f0_ghz:.5f} GHz")

                    print(f"    Measurement ... ", end="", flush=True)
                    result, com_retry = _measurement_solve(
                        project, param_names, x, f0_ghz, runner, msg_logger, sample_idx=i
                    )

                    if com_retry:
                        print(f"COM error, reconnecting...", flush=True)
                        try:
                            conn.reconnect()
                            project = conn.open_project(project_path)
                            result, _ = _measurement_solve(
                                project, param_names, x, f0_ghz, runner, msg_logger, sample_idx=i
                            )
                        except Exception:
                            result = {"solver_ok": False, "error": "COM unrecoverable", "f0_ghz": f0_ghz}

                    elapsed = time.perf_counter() - t0
                    result["x"] = x
                    result["f0_calibrated_ghz"] = f0_ghz
                    result["elapsed_s"] = elapsed

                    if result.get("solver_ok"):
                        print(f"OK ({elapsed:.0f}s)  Beta={result.get('coupling_beta',np.nan):.3f}")
                    else:
                        print(f"FAIL ({elapsed:.0f}s)")

                _excel_append(xlsx_path, _make_row(i, batch, param_names, output_entries, result))
                all_results.append(result)
                eval_count += 1
                completed_indices.add(i)
                _save_checkpoint(checkpoint_path, completed_indices, failed_permanent)

                # ── Post-sample graceful reset ──────────────────────
                if retry_handler is not None:
                    try:
                        retry_handler.force_reset()
                    except Exception:
                        pass

            # ── Convergence check ─────────────────────────────────────
            current_means = {}
            for out in output_entries:
                key = out["name"]
                vals = [r.get(key, np.nan) for r in all_results if r["solver_ok"]]
                vals = [v for v in vals if np.isfinite(v)]
                if len(vals) >= 3:
                    current_means[key] = float(np.mean(vals))

            if _check_convergence(prev_means, current_means, conv_rtol):
                print(f"\n  CONVERGED — all output means stable to <{conv_rtol*100:.0f}%")
                break

            prev_means = current_means
            current_n += batch_size
            if current_n > max_samples:
                print(f"\n  Max samples ({max_samples}) reached — stopping.")
                break

    # ── Cleanup ──────────────────────────────────────────────────────
    # Close retry handler first (cleans up any connections created during escalation)
    if retry_handler is not None:
        try:
            retry_handler.close_all(force=True)
        except Exception:
            pass
    else:
        try:
            if project is not None:
                project.close(save=False)
        except Exception:
            pass
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass

    # Remove stale lock file if present
    from cst_optimization.core.cleanup import remove_lock_file
    prj_dir = _os.path.splitext(project_path)[0]
    remove_lock_file(prj_dir)

    # ── Final summary ─────────────────────────────────────────────────
    valid = [r for r in all_results if r["solver_ok"]]
    if not valid:
        print("\nNo successful evaluations."); return

    print(f"\n{'='*80}")
    print(f"  TOLERANCE RESULTS  ({len(valid)}/{len(all_results)} OK, {batch} batches)")
    print(f"{'='*80}")

    for out in output_entries:
        key = out["name"]
        vals = np.array([r.get(key, np.nan) for r in valid])
        vals = vals[np.isfinite(vals)]
        if len(vals) < 3:
            print(f"\n  {key}: insufficient data"); continue
        m, s = np.mean(vals), np.std(vals, ddof=1)
        print(f"\n  {key:20s}  {out.get('description','')}")
        print(f"    Mean  = {m:.6g}  ± {s:.4g}  (1σ)")
        print(f"    Range = [{np.min(vals):.6g}, {np.max(vals):.6g}]")
        print(f"    CV    = {s/abs(m)*100:.2f}%")

    print(f"\nExcel: {xlsx_path}")
    print("Done.")


def _make_row(idx, batch, param_names, output_entries, result) -> list:
    row = [idx + 1, batch]
    x = result.get("x", np.zeros(len(param_names)))
    for j in range(len(param_names)):
        row.append(float(x[j]))
    row.append(result.get("f0_calibrated_ghz", np.nan))
    for o in output_entries:
        val = result.get(o["name"], np.nan)
        row.append(float(val) if np.isfinite(val) else "NaN")
    row.append(result.get("solver_ok", False))
    row.append(result.get("elapsed_s", 0.0))
    row.append(result.get("error", ""))
    row.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    return row


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Fatal error: {exc}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
    sys.exit(0)
