"""Analyse tolerance results to rank parameter sensitivity per output.

Reads ``tolerance_analysis.xlsx`` (or any Excel produced by
``tolerance_analysis.py``), extracts the (X, F) data from successful
evaluations, and runs three sensitivity methods to determine which
parameters drive each output's variation.

Usage
-----
::

    .venv\\Scripts\\python examples\\analyze_tolerance_sensitivity.py

Output
------
Per output: a ranked table with Sobol S1/ST, Spearman ρ, and linear β.
Parameters with the highest total-effect Sobol index (ST) are the ones
where tightening tolerance yields the greatest benefit.
"""

import sys
import os as _os
from datetime import datetime
_PROJECT_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_SRC_DIR = _os.path.join(_PROJECT_ROOT, "src")
if _SRC_DIR not in sys.path:
    sys.path.insert(0, _SRC_DIR)

import yaml
import numpy as np

from cst_optimization.optimization.in_situ_sensitivity import InSituSensitivity
from cst_optimization.optimization.base import OptimizationResult

CONFIG_PATH = _os.path.join(_PROJECT_ROOT, "config", "default.yaml")
DEFAULT_XLSX = "D:/Results/tolerance_analysis.xlsx"

# Tee output: write to both stdout and a .txt file
_txt_fh = None


def _tee(*args, **kwargs) -> None:
    """Print to stdout and also to the open .txt file."""
    import builtins
    builtins.print(*args, **kwargs)
    if _txt_fh is not None:
        builtins.print(*args, **kwargs, file=_txt_fh, flush=True)


def main() -> None:
    cfg = yaml.safe_load(open(CONFIG_PATH, "r", encoding="utf-8"))
    log_dir = cfg.get("logging", {}).get("output_dir", "D:/Results")
    xlsx_path = _os.path.join(log_dir, "tolerance_analysis.xlsx")

    # Open txt output
    global _txt_fh
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    txt_path = _os.path.join(log_dir, f"tolerance_sensitivity_{ts}.txt")
    _txt_fh = open(txt_path, "w", encoding="utf-8")
    _tee(f"Output file: {txt_path}\n")

    if not _os.path.exists(xlsx_path):
        _tee(f"Excel not found: {xlsx_path}")
        _tee(f"Run tolerance_analysis.py first, or provide a path.")
        return

    # ── Read Excel ───────────────────────────────────────────────────
    try:
        import openpyxl
    except ImportError:
        _tee("openpyxl required: pip install openpyxl"); return

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Read header row to find column indices
    headers_raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [str(h) if h is not None else "" for h in headers_raw]

    # Identify parameter columns (start with "x_" or appear in tolerance params)
    tc = cfg.get("tolerance", {})
    enabled_params = [p["name"] for p in tc.get("parameters", []) if p.get("enabled")]
    param_names = enabled_params if enabled_params else [
        h[2:] for h in headers if h.startswith("x_")
    ]

    output_names = [o["name"] for o in tc.get("outputs", []) if o.get("enabled")]
    if not output_names:
        output_names = ["f0_ghz", "q_loaded", "coupling_beta", "q0",
                        "e_peak", "s11_db", "p_input_mw"]

    # Build column index maps
    param_cols = {}
    for name in param_names:
        for j, h in enumerate(headers):
            if h == f"x_{name}" or h == name:
                param_cols[name] = j
                break

    output_cols = {}
    for name in output_names:
        for j, h in enumerate(headers):
            if h == name:
                output_cols[name] = j
                break

    # solver_ok column
    ok_col = headers.index("solver_ok") if "solver_ok" in headers else -1

    # Read data rows
    X_list = []
    F_lists = {name: [] for name in output_names}
    n_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if ok_col >= 0 and row[ok_col] != True and row[ok_col] != "True":
            continue  # skip failed rows
        try:
            x_row = [float(row[param_cols[n]]) for n in param_names if n in param_cols]
        except (ValueError, TypeError):
            continue
        X_list.append(np.array(x_row))
        for name in output_names:
            if name in output_cols:
                try:
                    F_lists[name].append(float(row[output_cols[name]]))
                except (ValueError, TypeError):
                    F_lists[name].append(np.nan)
        n_rows += 1

    wb.close()

    if n_rows < 10:
        _tee(f"Only {n_rows} valid rows — need at least 10 for correlation, 30 for gp_sobol.")
        return

    # Keep only params that actually appear in the data
    actual_params = [n for n in param_names if n in param_cols]
    _tee(f"Data: {n_rows} rows × {len(actual_params)} params × {len(output_names)} outputs\n")

    # ── Sensitivity per output ───────────────────────────────────────
    for out_name in output_names:
        y_raw = np.array(F_lists[out_name])
        valid = np.isfinite(y_raw)
        if valid.sum() < 10:
            _tee(f"\n[{out_name}] <10 valid points — skipping\n")
            continue

        X_valid = np.array([X_list[i] for i in range(n_rows) if valid[i]])
        y_valid = y_raw[valid]
        n_valid = len(y_valid)

        result = OptimizationResult(
            x_opt=X_valid[0], f_opt=np.array([y_valid[0]]),
            history_x=[X_valid[i] for i in range(n_valid)],
            history_f=[np.array([y_valid[i]]) for i in range(n_valid)],
            n_evaluations=n_valid,
        )

        analyzer = InSituSensitivity(result, actual_params, seed=42)

        _tee(f"{'=' * 70}")
        _tee(f"  {out_name}  (n={n_valid})")
        _tee(f"{'=' * 70}")

        # --- GP-Sobol ---
        if n_valid >= 30:
            try:
                r_sobol = analyzer.analyze("gp_sobol", n_gp_base_samples=512)
                print_ranking("GP-Sobol ST (total effect)", r_sobol.sobol_st, actual_params)
            except Exception as exc:
                _tee(f"  gp_sobol: FAILED — {exc}")

        # --- Spearman ---
        try:
            r_corr = analyzer.analyze("correlation")
            print_ranking("Spearman ρ", r_corr.spearman, actual_params)
        except Exception as exc:
            _tee(f"  correlation: FAILED — {exc}")

        # --- Linear ---
        try:
            r_lin = analyzer.analyze("linear")
            print_ranking("Linear β (standardised)", r_lin.linear_betas, actual_params)
        except Exception as exc:
            _tee(f"  linear: FAILED — {exc}")

        # --- Summary recommendation ---
        if n_valid >= 30:
            try:
                report = analyzer.analyze("gp_sobol", n_gp_base_samples=512)
                _tee(f"\n  RECOMMENDATION: {report.recommendation}")
            except Exception:
                pass
        _tee()

    _tee("Done.")
    if _txt_fh is not None:
        _txt_fh.close()


def print_ranking(title: str, scores: dict[str, float], param_names: list[str]) -> None:
    """Print a ranked table of parameter sensitivity scores."""
    ranked = sorted(scores.items(), key=lambda kv: abs(kv[1]), reverse=True)
    _tee(f"\n  {title}:")
    _tee(f"  {'RANK':<5} {'PARAMETER':<22} {'VALUE':>8}  IMPACT")
    _tee(f"  {'─'*5} {'─'*22} {'─'*8}  {'─'*15}")

    for rank, (name, val) in enumerate(ranked, 1):
        if abs(val) > 0.5:
            impact = "HIGH — prioritise tight tolerance"
        elif abs(val) > 0.2:
            impact = "MEDIUM"
        elif abs(val) > 0.05:
            impact = "LOW — can relax"
        else:
            impact = "NEGLIGIBLE"
        _tee(f"  {rank:<5} {name:<22} {val:>+8.4f}  {impact}")


if __name__ == "__main__":
    main()
