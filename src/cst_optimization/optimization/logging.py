"""Excel-based optimisation logger — dual-sheet transparency.

Writes every simulation evaluation and every optimizer decision to a
single ``.xlsx`` workbook.  CST 2026 refactor: incremental append via
``load_workbook`` + dirty-tracking instead of O(n²) full-rewrite.
"""

from __future__ import annotations

import logging
import os
import threading
from datetime import datetime
from typing import Any

import numpy as np

_logger = logging.getLogger(__name__)


class OptimizationLogger:
    """Dual-sheet Excel logger with incremental I/O.

    Parameters
    ----------
    filepath : str
        Path to the output ``.xlsx`` file.
    auto_flush_interval : int
        Save to disk every *N* evaluations.  Default 5.
    """

    def __init__(self, filepath: str, auto_flush_interval: int = 5) -> None:
        self._filepath = filepath
        self._auto_flush = auto_flush_interval
        self._eval_rows: list[dict[str, Any]] = []
        self._decision_rows: list[dict[str, Any]] = []
        self._dirty = False
        self._lock = threading.Lock()

    # ------------------------------------------------------------------
    # Public logging API
    # ------------------------------------------------------------------

    def log_evaluation(
        self,
        iteration: int,
        x: np.ndarray,
        param_names: list[str],
        physics: dict[str, float],
        objective_values: dict[str, float],
        solver_ok: bool = True,
        error: str = "",
        elapsed_s: float = 0.0,
    ) -> None:
        """Record one completed CST simulation."""
        row: dict[str, Any] = {"iter": iteration}
        for name, val in zip(param_names, x):
            row[f"x_{name}"] = float(val)
        for key, val in physics.items():
            row[key] = float(val) if np.isfinite(val) else "NaN"
        for key, val in objective_values.items():
            row[f"obj_{key}"] = float(val) if np.isfinite(val) else "NaN"
        row["solver_ok"] = solver_ok
        row["error"] = error
        row["elapsed_s"] = round(elapsed_s, 1)
        row["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with self._lock:
            self._eval_rows.append(row)
            self._dirty = True

        if len(self._eval_rows) % self._auto_flush == 0:
            self.flush()

    def log_decision(
        self,
        iteration: int,
        x_proposed: np.ndarray,
        param_names: list[str],
        acquisition_value: float,
        y_best_so_far: float,
        gp_params: dict[str, Any] | None = None,
    ) -> None:
        """Record one optimizer decision."""
        row: dict[str, Any] = {"iter": iteration}
        for name, val in zip(param_names, x_proposed):
            row[f"x_{name}_proposed"] = float(val)
        row["acquisition"] = float(acquisition_value)
        row["y_best_so_far"] = float(y_best_so_far)

        if gp_params:
            for k, v in gp_params.items():
                row[f"gp_{k}"] = str(v) if isinstance(v, (list, np.ndarray)) else v

        row["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with self._lock:
            self._decision_rows.append(row)
            self._dirty = True

    # ------------------------------------------------------------------
    # Persistence
    # ------------------------------------------------------------------

    def flush(self) -> None:
        """Write buffered rows to disk (incremental when possible)."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel logging.  "
                "Install with: pip install openpyxl"
            )

        with self._lock:
            if not self._dirty:
                return

            try:
                os.makedirs(os.path.dirname(self._filepath) or ".", exist_ok=True)

                # Incremental: load existing workbook if available,
                # otherwise create a new one
                if os.path.isfile(self._filepath):
                    wb = openpyxl.load_workbook(self._filepath)
                else:
                    wb = openpyxl.Workbook()

                self._write_eval_sheet(wb)
                self._write_decision_sheet(wb)
                wb.save(self._filepath)
                self._dirty = False
            except Exception:
                _logger.warning(
                    "Failed to flush optimization log to %s",
                    self._filepath, exc_info=True,
                )

    def save(self) -> None:
        """Final save with formatting."""
        try:
            import openpyxl
        except ImportError:
            return

        with self._lock:
            try:
                os.makedirs(os.path.dirname(self._filepath) or ".", exist_ok=True)
                wb = openpyxl.Workbook()
                self._write_eval_sheet(wb)
                self._write_decision_sheet(wb)
                wb.save(self._filepath)
                self._dirty = False
            except Exception:
                _logger.warning(
                    "Failed to save final optimization log", exc_info=True,
                )

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _write_eval_sheet(self, wb: Any) -> None:
        if not self._eval_rows:
            return

        import openpyxl

        ws = wb.active
        if ws.title != "Evaluations" and not self._is_virgin_sheet(ws):
            ws = wb.create_sheet("Evaluations")
        else:
            ws.title = "Evaluations"
            ws.delete_rows(1, ws.max_row)

        headers = list(self._eval_rows[0].keys())
        ws.append(headers)
        for row in self._eval_rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"

    def _write_decision_sheet(self, wb: Any) -> None:
        if not self._decision_rows:
            return

        import openpyxl

        if "Decisions" in wb.sheetnames:
            ws = wb["Decisions"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Decisions")

        headers = list(self._decision_rows[0].keys())
        ws.append(headers)
        for row in self._decision_rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"

    @staticmethod
    def _is_virgin_sheet(ws: Any) -> bool:
        try:
            return ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None
        except Exception:
            return True

    # ------------------------------------------------------------------
    # Properties
    # ------------------------------------------------------------------

    @property
    def filepath(self) -> str:
        return self._filepath

    @property
    def n_evaluations(self) -> int:
        return len(self._eval_rows)

    @property
    def n_decisions(self) -> int:
        return len(self._decision_rows)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def extract_gp_params(gp_model: Any) -> dict[str, Any]:
    """Extract readable GP hyperparameters from a fitted ``GaussianProcessRegressor``."""
    kernel = gp_model.kernel_
    params: dict[str, Any] = {"kernel": str(kernel)}

    try:
        if hasattr(kernel, "k1") and hasattr(kernel.k1, "k2"):
            signal_var = kernel.k1.k1.constant_value
            length_scales = kernel.k1.k2.length_scale
            noise = kernel.k2.noise_level
            params["signal_variance"] = float(signal_var)
            params["length_scales"] = (
                list(length_scales)
                if hasattr(length_scales, "__len__")
                else float(length_scales)
            )
            params["noise_level"] = float(noise)
    except Exception:
        pass

    try:
        params["log_marginal_likelihood"] = float(
            gp_model.log_marginal_likelihood_value_
        )
    except Exception:
        pass

    return params
