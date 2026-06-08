"""Unified diagnostics module — error types, CST message capture, and optimisation logging.

This module consolidates three previously separate modules:
- ``core/errors.py`` — structured exception hierarchy
- ``core/messages.py`` — CST Message Window capture and persistence
- ``optimization/logging.py`` — Excel-based optimisation logger

All components have independent enable/disable switches so callers only pay for
what they need.
"""

from __future__ import annotations

import logging
import os
import threading
from datetime import datetime
from typing import Any, TYPE_CHECKING

import numpy as np

if TYPE_CHECKING:
    from .core.project import CSTProject

# ============================================================================
# Exception hierarchy (from core/errors.py)
# ============================================================================


class CSTError(Exception):
    """Base exception for all CST interaction errors."""


class CSTConnectionLostError(CSTError):
    """DesignEnvironment connection was lost (COM dead, process killed).

    The orchestrator should attempt a full reconnect + single retry
    when this is raised.
    """


# -- Solver errors -----------------------------------------------------------


class SolverError(CSTError):
    """Base class for solver execution failures.

    Attributes
    ----------
    elapsed_s : float | None
        Wall-clock time spent in the solver before failure.
    """

    def __init__(self, *args: object, elapsed_s: float | None = None) -> None:
        super().__init__(*args)
        self.elapsed_s = elapsed_s


class SolverTimeoutError(SolverError):
    """Solver exceeded the configured timeout."""


class SolverMeshError(SolverError):
    """Mesh generation failed (distortion, topology intersection, etc.).

    The orchestrator should log this as a geometry issue and skip the
    current sample rather than attempting a retry.
    """


class SolverConvergenceError(SolverError):
    """Solver failed to converge within the allowed number of iterations."""


class SolverCOMError(SolverError):
    """COM layer communication failure — the connection to CST is broken.

    The orchestrator should attempt a reconnect + single retry.
    """


class SolverStagnationError(SolverError):
    """Solver progress stalled beyond the stagnation timeout."""


# ============================================================================
# CST Message Logger (from core/messages.py)
# ============================================================================

_logger = logging.getLogger(__name__)


class MessageLogger:
    """Capture CST Message Window content and persist to disk.

    Parameters
    ----------
    output_dir : str
        Directory for message log files.
    enabled : bool
        If ``False``, all operations are no-ops (production bypass).
    """

    def __init__(self, output_dir: str = "", enabled: bool = True) -> None:
        self._output_dir = output_dir
        self._enabled = enabled
        self._buffer: list[str] = []

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def capture(self, project: CSTProject) -> str:
        """Read current messages from the CST project and append to buffer."""
        if not self._enabled:
            return ""

        try:
            raw = project.get_messages()
        except Exception as exc:
            raw = f"[MessageLogger] get_messages() raised: {exc}"

        text = self._normalize(raw)
        if text:
            self._buffer.append(text)
        return text

    def write(self, label: str = "", iteration: int = 0) -> str | None:
        """Flush buffered messages to a timestamped file."""
        if not self._enabled or not self._buffer:
            return None

        try:
            os.makedirs(self._output_dir, exist_ok=True)
        except OSError:
            _logger.warning(
                "Cannot create message output dir: %s", self._output_dir
            )
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"msg_{label}_iter{iteration:04d}_{ts}.txt"
        fpath = os.path.join(self._output_dir, fname)

        try:
            with open(fpath, "w", encoding="utf-8") as fh:
                fh.write(f"CST Messages — {label} — iteration {iteration}\n")
                fh.write(f"Captured: {datetime.now().isoformat()}\n")
                fh.write("=" * 72 + "\n\n")
                for i, msg in enumerate(self._buffer):
                    fh.write(f"--- Capture {i + 1} ---\n")
                    fh.write(msg)
                    fh.write("\n\n")
        except OSError:
            _logger.warning("Failed to write message log: %s", fpath, exc_info=True)
            return None

        self._buffer.clear()
        return fpath

    def write_now(
        self, text: str, label: str = "", iteration: int = 0
    ) -> str | None:
        """Directly write a string to a message file (bypasses buffer)."""
        if not self._enabled or not text.strip():
            return None

        try:
            os.makedirs(self._output_dir, exist_ok=True)
        except OSError:
            _logger.warning(
                "Cannot create message output dir: %s", self._output_dir
            )
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"msg_{label}_iter{iteration:04d}_{ts}.txt"
        fpath = os.path.join(self._output_dir, fname)

        try:
            with open(fpath, "w", encoding="utf-8") as fh:
                fh.write(f"CST Messages — {label} — iteration {iteration}\n")
                fh.write(f"Captured: {datetime.now().isoformat()}\n")
                fh.write("=" * 72 + "\n\n")
                fh.write(text)
                fh.write("\n")
        except OSError:
            _logger.warning(
                "Failed to write message log: %s", fpath, exc_info=True
            )
            return None

        return fpath

    def clear(self) -> None:
        """Discard all buffered messages."""
        self._buffer.clear()

    @property
    def is_empty(self) -> bool:
        """Return ``True`` if the buffer is empty."""
        return len(self._buffer) == 0

    # Known fatal VBA history replay failure patterns.
    _HISTORY_FAILURE_PATTERNS: tuple[str, ...] = (
        "history update failed",
        "not positioned at the very last entry",
    )

    def has_history_failure(self) -> bool:
        """Check buffered messages for VBA history replay failures."""
        if not self._buffer:
            return False
        for msg in self._buffer:
            msg_lower = msg.lower()
            if any(p in msg_lower for p in self._HISTORY_FAILURE_PATTERNS):
                return True
        return False

    @staticmethod
    def _normalize(raw: object) -> str:
        """Convert the opaque ``get_messages()`` return to a string."""
        if raw is None:
            return ""
        if isinstance(raw, str):
            return raw.strip()
        if isinstance(raw, (list, tuple)):
            return "\n".join(str(item) for item in raw).strip()
        if isinstance(raw, dict):
            return "\n".join(f"{k}: {v}" for k, v in raw.items()).strip()
        try:
            return str(raw).strip()
        except Exception:
            return repr(raw).strip()


# ============================================================================
# Optimisation Excel Logger (from optimization/logging.py)
# ============================================================================


class OptimizationLogger:
    """Dual-sheet Excel logger with incremental I/O.

    Parameters
    ----------
    filepath : str
        Path to the output ``.xlsx`` file.
    auto_flush_interval : int
        Save to disk every *N* evaluations.  Default 5.
    """

    def __init__(self, filepath: str, auto_flush_interval: int = 5) -> None:
        self._filepath = filepath
        self._auto_flush = auto_flush_interval
        self._eval_rows: list[dict[str, Any]] = []
        self._decision_rows: list[dict[str, Any]] = []
        self._dirty = False
        self._lock = threading.Lock()

    def log_evaluation(
        self,
        iteration: int,
        x: np.ndarray,
        param_names: list[str],
        physics: dict[str, float],
        objective_values: dict[str, float],
        solver_ok: bool = True,
        error: str = "",
        elapsed_s: float = 0.0,
    ) -> None:
        """Record one completed CST simulation."""
        row: dict[str, Any] = {"iter": iteration}
        for name, val in zip(param_names, x):
            row[f"x_{name}"] = float(val)
        for key, val in physics.items():
            row[key] = float(val) if np.isfinite(val) else "NaN"
        for key, val in objective_values.items():
            row[f"obj_{key}"] = float(val) if np.isfinite(val) else "NaN"
        row["solver_ok"] = solver_ok
        row["error"] = error
        row["elapsed_s"] = round(elapsed_s, 1)
        row["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with self._lock:
            self._eval_rows.append(row)
            self._dirty = True

        if len(self._eval_rows) % self._auto_flush == 0:
            self.flush()

    def log_decision(
        self,
        iteration: int,
        x_proposed: np.ndarray,
        param_names: list[str],
        acquisition_value: float,
        y_best_so_far: float,
        gp_params: dict[str, Any] | None = None,
    ) -> None:
        """Record one optimizer decision."""
        row: dict[str, Any] = {"iter": iteration}
        for name, val in zip(param_names, x_proposed):
            row[f"x_{name}_proposed"] = float(val)
        row["acquisition"] = float(acquisition_value)
        row["y_best_so_far"] = float(y_best_so_far)

        if gp_params:
            for k, v in gp_params.items():
                row[f"gp_{k}"] = str(v) if isinstance(v, (list, np.ndarray)) else v

        row["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with self._lock:
            self._decision_rows.append(row)
            self._dirty = True

    def flush(self) -> None:
        """Write buffered rows to disk (incremental when possible)."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel logging.  "
                "Install with: pip install openpyxl"
            )

        with self._lock:
            if not self._dirty:
                return

            try:
                os.makedirs(os.path.dirname(self._filepath) or ".", exist_ok=True)
                if os.path.isfile(self._filepath):
                    wb = openpyxl.load_workbook(self._filepath)
                else:
                    wb = openpyxl.Workbook()

                self._write_eval_sheet(wb)
                self._write_decision_sheet(wb)
                wb.save(self._filepath)
                self._dirty = False
            except Exception:
                _logger.warning(
                    "Failed to flush optimization log to %s",
                    self._filepath, exc_info=True,
                )

    def save(self) -> None:
        """Final save with formatting."""
        try:
            import openpyxl
        except ImportError:
            return

        with self._lock:
            try:
                os.makedirs(os.path.dirname(self._filepath) or ".", exist_ok=True)
                wb = openpyxl.Workbook()
                self._write_eval_sheet(wb)
                self._write_decision_sheet(wb)
                wb.save(self._filepath)
                self._dirty = False
            except Exception:
                _logger.warning(
                    "Failed to save final optimization log", exc_info=True,
                )

    def _write_eval_sheet(self, wb: Any) -> None:
        if not self._eval_rows:
            return
        import openpyxl
        ws = wb.active
        if ws.title != "Evaluations" and not self._is_virgin_sheet(ws):
            ws = wb.create_sheet("Evaluations")
        else:
            ws.title = "Evaluations"
            ws.delete_rows(1, ws.max_row)
        headers = list(self._eval_rows[0].keys())
        ws.append(headers)
        for row in self._eval_rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"

    def _write_decision_sheet(self, wb: Any) -> None:
        if not self._decision_rows:
            return
        import openpyxl
        if "Decisions" in wb.sheetnames:
            ws = wb["Decisions"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Decisions")
        headers = list(self._decision_rows[0].keys())
        ws.append(headers)
        for row in self._decision_rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"

    @staticmethod
    def _is_virgin_sheet(ws: Any) -> bool:
        try:
            return ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None
        except Exception:
            return True

    @property
    def filepath(self) -> str:
        return self._filepath

    @property
    def n_evaluations(self) -> int:
        return len(self._eval_rows)

    @property
    def n_decisions(self) -> int:
        return len(self._decision_rows)


def extract_gp_params(gp_model: Any) -> dict[str, Any]:
    """Extract readable GP hyperparameters from a fitted ``GaussianProcessRegressor``."""
    kernel = gp_model.kernel_
    params: dict[str, Any] = {"kernel": str(kernel)}
    try:
        if hasattr(kernel, "k1") and hasattr(kernel.k1, "k2"):
            signal_var = kernel.k1.k1.constant_value
            length_scales = kernel.k1.k2.length_scale
            noise = kernel.k2.noise_level
            params["signal_variance"] = float(signal_var)
            params["length_scales"] = (
                list(length_scales)
                if hasattr(length_scales, "__len__")
                else float(length_scales)
            )
            params["noise_level"] = float(noise)
    except Exception:
        pass
    try:
        params["log_marginal_likelihood"] = float(
            gp_model.log_marginal_likelihood_value_
        )
    except Exception:
        pass
    return params
