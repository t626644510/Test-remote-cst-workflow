"""One-shot script: migrate existing Excel optimisation log to JSONL raw database.

Usage::

    python examples/migrate_excel_to_db.py

Reads ``D:/Results/optimization_log.xlsx`` and writes
``D:/Results/raw_evaluations.jsonl``.
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime

# Ensure we can import the project package
_project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_src = os.path.join(_project_root, "src")
if _src not in sys.path:
    sys.path.insert(0, _src)

EXCEL_PATH = "D:/Results/optimization_log.xlsx"
OUT_PATH = "D:/Results/raw_evaluations.jsonl"

# ── Column mapping ──────────────────────────────────────────────────────
# Parameter columns in the Excel (prefix "x_" stripped → param name)
PARAM_NAMES = [
    "selfangle1", "selfangle2", "inner_angle", "inner_angle3",
    "FolkHeight", "FolkHeight2", "UpperHeight1", "UpperHeight2",
    "DownHeight1", "DownHeight2", "Lin2", "inner_r2", "Lin", "inner_r",
]

# Raw objective columns (exact header names in the Excel)
RAW_OBJ_NAMES = [
    "z_longitudinal", "z_transverse",
    "antenna_absorption", "antenna_absorption_db",
]


def main() -> None:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required: pip install openpyxl")
        sys.exit(1)

    if not os.path.isfile(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Evaluations"]

    # Read header row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {h: idx for idx, h in enumerate(headers) if h is not None}

    records_written = 0
    os.makedirs(os.path.dirname(OUT_PATH) or ".", exist_ok=True)

    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

            iter_val = _cell(row_vals, col_map, "iter", 0)
            solver_ok = bool(_cell(row_vals, col_map, "solver_ok", True))
            error = _cell(row_vals, col_map, "error", "") or ""
            timestamp = _cell(row_vals, col_map, "timestamp",
                             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            # Build params dict
            params = {}
            for pn in PARAM_NAMES:
                col_name = f"x_{pn}"
                if col_name in col_map:
                    v = row_vals[col_map[col_name]]
                    params[pn] = float(v) if v is not None else None

            # Build raw dict
            raw = {}
            for on_ in RAW_OBJ_NAMES:
                if on_ in col_map:
                    v = row_vals[col_map[on_]]
                    raw[on_] = float(v) if v is not None else None

            record = {
                "iter": int(iter_val) if iter_val is not None else 0,
                "timestamp": str(timestamp),
                "params": params,
                "raw": raw,
                "solver_ok": solver_ok,
                "error": error if error else None,
            }
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")
            records_written += 1

    print(f"Migrated {records_written} records")
    print(f"  Source: {EXCEL_PATH}")
    print(f"  Output: {OUT_PATH}")


def _cell(row: list, col_map: dict, key: str, default: object) -> object:
    idx = col_map.get(key)
    if idx is None:
        return default
    return row[idx]


if __name__ == "__main__":
    main()
